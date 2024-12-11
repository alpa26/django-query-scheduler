import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import dramatiq
import requests
from django.utils import timezone
from .models import Task, TaskResult
import sqlite3


def enter_data_in_table(task, response):
    print(f"table method")
    directory = 'excel_tables'
    file_path = os.path.join(os.path.dirname(__file__), '..', directory, task.table_name)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        print(f"found table {file_path}")
    except FileNotFoundError as err:
        print(f"not found table {file_path}")
        workbook = Workbook()
        sheet = workbook.active
        headers = ["Name", "Type", "Resource", "Request", "Responce", "ScheduleTime"]
        sheet.append(headers)


    row = [
        task.name,
        task.task_type,
        task.resource,
        task.request,
        response,
        task.schedule_time.strftime('%Y-%m-%d %H:%M:%S')
    ]
    print(f"{task.name},{task.task_type},{task.resource},{task.request},{response}, {task.schedule_time.strftime('%Y-%m-%d %H:%M:%S')}")


    sheet.append(row)

    workbook.save(file_path)


@dramatiq.actor
def run_api_task(task_id, attempt=1):
    task = Task.objects.get(id=task_id)
    if (not task.is_active):
        return
    print("found api")
    try:
        response = requests.get(task.request)
        response.raise_for_status()
        result = response.text
        status = 'success'
        task.last_run = timezone.now()
        task.save()
        if(task.table_name):
            enter_data_in_table(task, result)
        TaskResult.objects.create(task=task, result=result, status=status)
        # Если задача вызывается с параметром attempt == -1 то задача выполняется 1 раз
        if attempt != -1:
            if task.repeat_interval == 'hourly':
                task.next_run = task.last_run + timezone.timedelta(days=1)
                delay = (task.next_run - timezone.now()).total_seconds() * 1000
                run_api_task.send_with_options(args=(task_id,), delay=int(delay))
            if task.repeat_interval == 'daily':
                task.next_run = task.last_run + timezone.timedelta(days=1)
                delay = (task.next_run - timezone.now()).total_seconds() * 1000
                run_api_task.send_with_options(args=(task_id,), delay=int(delay))
            if task.repeat_interval == 'weekly':
                task.next_run = task.last_run + timezone.timedelta(days=7)
                delay = (task.next_run - timezone.now()).total_seconds() * 1000
                run_api_task.send_with_options(args=(task_id,), delay=int(delay))

    except requests.exceptions.RequestException as e:
        if (attempt == -1):
            attempt = 1
        result = str(e)
        status = 'error'
        if (attempt < 4):
            TaskResult.objects.create(task=task, result=result, status=status)
            run_api_task.send_with_options(args=(task_id, attempt + 1,), delay=90000)


@dramatiq.actor
def run_db_task(task_id, attempt=1):
    task = Task.objects.get(id=task_id)
    if (not task.is_active):
        return
    print("try")
    try:
        db_path = os.path.join(os.path.dirname(__file__), '..', 'db', f"{task.resource}.sqlite3")
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        conn = sqlite3.connect(db_path)

        cursor = conn.cursor()
        cursor.execute(task.request)
        result = cursor.fetchall()

        conn.commit()
        conn.close()

        status = 'success'
        task.last_run = timezone.now()
        task.save()
        if isinstance(result, list):
            result_str = ', '.join(map(str, result))
        else:
            result_str = result
        if(task.table_name):
            enter_data_in_table(task, result_str)
        TaskResult.objects.create(task=task, result=result, status=status)
        # Если задача вызывается с параметром attempt == -1 то задача выполняется 1 раз
        if attempt != -1:
            if task.repeat_interval == 'hourly':
                task.next_run = task.last_run + timezone.timedelta(hours=1)
                delay = (task.next_run - timezone.now()).total_seconds() * 1000
                run_db_task.send_with_options(args=(task_id,), delay=int(delay))
            if task.repeat_interval == 'daily':
                task.next_run = task.last_run + timezone.timedelta(days=1)
                delay = (task.next_run - timezone.now()).total_seconds() * 1000
                run_db_task.send_with_options(args=(task_id,), delay=int(delay))
            if task.repeat_interval == 'weekly':
                task.next_run = task.last_run + timezone.timedelta(days=1)
                delay = (task.next_run - timezone.now()).total_seconds() * 1000
                run_db_task.send_with_options(args=(task_id,), delay=int(delay))
    except Exception as e:
        if (attempt == -1):
            attempt = 1
        result = str(e)
        status = 'error'
        if (attempt < 4):
            TaskResult.objects.create(task=task, result=result, status=status)
            run_db_task.send_with_options(args=(task_id, attempt + 1,), delay=10000)